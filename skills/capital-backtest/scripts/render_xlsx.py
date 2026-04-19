"""多标签 xlsx 汇总：每个 bucket 一个 monthly sheet、一个 trade log sheet，外加 summary sheet。

用 openpyxl。只依赖 pandas + openpyxl 这两个常见包。
"""

from __future__ import annotations

from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _autofit(ws, df: pd.DataFrame, max_width: int = 42) -> None:
    for idx, col in enumerate(df.columns, start=1):
        sample = df[col].head(200)
        width = max([len(str(col))] + [len(str(s)) for s in sample])
        width = min(max(width + 2, 8), max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _color_pnl(ws, df: pd.DataFrame, pnl_col: str = "realized_pnl") -> None:
    if pnl_col not in df.columns:
        return
    col_idx = list(df.columns).index(pnl_col) + 1
    fill_up = PatternFill("solid", fgColor="DFF0D8")
    fill_dn = PatternFill("solid", fgColor="F2DEDE")
    for r in range(2, len(df) + 2):
        v = df.iloc[r - 2][pnl_col]
        if pd.isna(v):
            continue
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = fill_up if v > 0 else fill_dn if v < 0 else PatternFill()


def write_report(
    out_path: str | Path,
    summary_rows: pd.DataFrame,
    metrics_rows: pd.DataFrame,
    monthlies: Mapping[str, pd.DataFrame],
    yearlies: Mapping[str, pd.DataFrame],
    trade_logs: Mapping[str, pd.DataFrame],
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_rows.to_excel(writer, sheet_name="summary", index=False)
        metrics_rows.to_excel(writer, sheet_name="metrics", index=False)

        for bucket, m in monthlies.items():
            name = _safe_sheet_name(f"{bucket}_monthly")
            m.to_excel(writer, sheet_name=name, index=False)

        for bucket, y in yearlies.items():
            name = _safe_sheet_name(f"{bucket}_yearly")
            y.to_excel(writer, sheet_name=name, index=False)

        for bucket, t in trade_logs.items():
            name = _safe_sheet_name(f"{bucket}_trades")
            # trade log 可能很大，截到 20k 行
            if len(t) > 20_000:
                t = t.head(20_000)
            t.to_excel(writer, sheet_name=name, index=False)

        # 格式化
        wb = writer.book
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            if ws.max_row == 0:
                continue
            # 表头加粗
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).font = Font(bold=True)
                ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            # 自适宽 + PnL 着色
            df_for_sheet = _find_sheet_df(ws_name, summary_rows, metrics_rows, monthlies, yearlies, trade_logs)
            if df_for_sheet is not None:
                _autofit(ws, df_for_sheet)
                if ws_name.endswith("_monthly") or ws_name.endswith("_yearly") or ws_name == "metrics":
                    _color_pnl(ws, df_for_sheet, "realized_pnl")
                if ws_name.endswith("_trades"):
                    _color_pnl(ws, df_for_sheet, "pnl")

    return out_path


def _safe_sheet_name(name: str) -> str:
    # Excel sheet name 上限 31 字符，去掉非法字符
    bad = set(r"[]:*?/\\")
    cleaned = "".join("_" if c in bad else c for c in name)
    return cleaned[:31]


def _find_sheet_df(
    name: str,
    summary_rows: pd.DataFrame,
    metrics_rows: pd.DataFrame,
    monthlies: Mapping[str, pd.DataFrame],
    yearlies: Mapping[str, pd.DataFrame],
    trade_logs: Mapping[str, pd.DataFrame],
):
    if name == "summary":
        return summary_rows
    if name == "metrics":
        return metrics_rows
    for bucket, df in monthlies.items():
        if name == _safe_sheet_name(f"{bucket}_monthly"):
            return df
    for bucket, df in yearlies.items():
        if name == _safe_sheet_name(f"{bucket}_yearly"):
            return df
    for bucket, df in trade_logs.items():
        if name == _safe_sheet_name(f"{bucket}_trades"):
            return df.head(20_000) if len(df) > 20_000 else df
    return None
